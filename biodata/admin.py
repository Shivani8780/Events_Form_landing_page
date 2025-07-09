import os
import logging
import sys
from django.contrib import admin
from django.http import HttpResponse
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
import requests
from PIL import Image as PILImage
from io import BytesIO
from django import forms
from django.utils.html import format_html
import io
import zipfile
import re
from .models import CandidateBiodata, GalleryImage
from ckeditor.widgets import CKEditorWidget
from datetime import datetime
from .models import AdvancePassBooking , AdvanceBookletBooking, StageRegistration


@admin.action(description='Export selected stage registrations to Excel')
def export_selected_to_excel(modeladmin, request, queryset):
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from io import BytesIO
    from django.http import HttpResponse
    import requests
    from PIL import Image as PILImage

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stage Registrations"

    # Define headers
    headers = ['Name of Candidate', 'DOB', 'Gender', 'Current City', 'Booklet Sr No', 'Maritial Status', 'Education', 'Job/Business', 'Hobbies', 'Partner Choice', 'Photograph']
    ws.append(headers)

    # Set column widths
    column_widths = [20, 15, 10, 20, 15, 15, 20, 20, 20, 20, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


    row_num = 2
    for obj in queryset:
        row = [
            obj.name_of_candidate,
            obj.dob if isinstance(obj.dob, str) else (obj.dob.strftime('%Y-%m-%d') if obj.dob else ''),
            obj.gender,
            obj.current_city,
            obj.booklet_sr_no,
            obj.maritial_status,
            obj.education,
            obj.job_business,
            obj.hobbies,
            obj.partner_choice,
        ]
        ws.append(row)
        row_num += 1

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=stage_registrations.xlsx'

    # Save workbook to response
    wb.save(response)
    return response

@admin.action(description='Download selected stage registration images as zip')
def download_selected_images(modeladmin, request, queryset):
    import zipfile
    import os
    from django.http import HttpResponse
    from io import BytesIO
    import re

    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for obj in queryset:
            if obj.photograph:
                try:
                    img_path = obj.photograph.path
                    ext = os.path.splitext(img_path)[1].lower()
                    if os.path.exists(img_path):
                        safe_name = re.sub(r'[^a-zA-Z0-9_-]', '_', obj.name_of_candidate.strip())
                        filename = f"{safe_name}{ext}"
                        with open(img_path, 'rb') as img_file:
                            img_data = img_file.read()
                        zip_file.writestr(filename, img_data)
                except Exception as e:
                    pass
    zip_buffer.seek(0)
    response = HttpResponse(zip_buffer, content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename=stage_registration_images.zip'
    return response

@admin.register(StageRegistration)
class StageRegistrationAdmin(admin.ModelAdmin):
    list_display = ['name_of_candidate', 'dob', 'gender', 'current_city', 'booklet_sr_no', 'maritial_status', 'education', 'job_business', 'hobbies', 'partner_choice', 'photograph_preview']
    search_fields = ['name_of_candidate', 'current_city', 'booklet_sr_no', 'maritial_status', 'education', 'job_business', 'partner_choice']
    readonly_fields = ['photograph_preview']

    actions = ['export_selected_to_excel', 'download_selected_images']

    export_selected_to_excel = export_selected_to_excel
    download_selected_images = download_selected_images

    def photograph_preview(self, obj):
        if obj.photograph:
            return format_html('<img src="{}" style="max-height: 100px; max-width: 100px;" />', obj.photograph.url)
        return "-"
    photograph_preview.short_description = 'Photograph Preview'

from .models import CandidateBiodata, GalleryImage, AdvancePassBooking, AdvanceBookletBooking


@admin.register(AdvanceBookletBooking)
class AdvanceBookletBookingAdmin(admin.ModelAdmin):
    list_display = ['name', 'city', 'whatsapp_number', 'email', 'girls_booklet_with', 'boys_booklet_with', 'courier_address', 'total_amount', 'payment_screenshot_preview', 'created_at']
    list_filter = ['created_at']
    search_fields = ['name', 'city', 'whatsapp_number', 'email']

    readonly_fields = ['payment_screenshot_preview']

    actions = ['export_selected_to_excel', 'export_payment_screenshots_zip']

    def payment_screenshot_preview(self, obj):
        if obj.payment_screenshot:
            return format_html(f'<img src="{obj.payment_screenshot.url}" style="max-height: 100px; max-width: 100px;" />')
        return "-"

    @admin.action(description='Export selected advance booklet bookings to Excel')
    def export_selected_to_excel(self, request, queryset):
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as OpenpyxlImage
        from io import BytesIO
        from django.http import HttpResponse
        import requests
        from PIL import Image as PILImage

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Advance Booklet Bookings"

        # Define headers
        headers = ['Name', 'City', 'Whatsapp Number', 'Email', 'Girls Booklet With', 'Boys Booklet With', 'Courier Address', 'Total Amount', 'Created At', 'Payment Screenshot']
        ws.append(headers)

        # Set column widths
        column_widths = [20, 20, 20, 30, 20, 20, 40, 15, 20, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        row_num = 2
        for obj in queryset:
            row = [
                obj.name,
                obj.city,
                obj.whatsapp_number,
                obj.email,
                obj.girls_booklet_with,
                obj.boys_booklet_with,
                obj.courier_address,
                obj.total_amount,
                obj.created_at.strftime('%Y-%m-%d %H:%M:%S') if obj.created_at else '',
                '',  # Placeholder for image
            ]
            ws.append(row)

            # Embed payment screenshot image if available
            if obj.payment_screenshot:
                try:
                    img_url = obj.payment_screenshot.url
                    if img_url.startswith('/'):
                        base_url = 'https://bhudevnetwork.pythonanywhere.com/'  # Replace with your actual domain or base URL
                        img_url = base_url + img_url
                    response_img = requests.get(img_url)
                    response_img.raise_for_status()
                    img_data = BytesIO(response_img.content)
                    pil_img = PILImage.open(img_data)
                    img_byte_arr = BytesIO()
                    pil_img.save(img_byte_arr, format='PNG')
                    img_byte_arr.seek(0)
                    img = OpenpyxlImage(img_byte_arr)
                    img.width = 80
                    img.height = 80
                    img.anchor = f"{get_column_letter(10)}{row_num}"
                    ws.add_image(img)
                    ws.row_dimensions[row_num].height = 60
                except Exception as e:
                    # Log error or pass silently
                    pass
            row_num += 1

        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=advance_booklet_bookings.xlsx'

        # Save workbook to response
        wb.save(response)
        return response

    @admin.action(description='Export payment screenshots of selected advance booklet bookings as ZIP')
    def export_payment_screenshots_zip(self, request, queryset):
        import zipfile
        import os
        from django.http import HttpResponse
        from io import BytesIO
        import re

        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for obj in queryset:
                if obj.payment_screenshot:
                    try:
                        img_path = obj.payment_screenshot.path
                        ext = os.path.splitext(img_path)[1].lower()
                        if os.path.exists(img_path):
                            # Sanitize name and whatsapp number for filename
                            safe_name = re.sub(r'[^a-zA-Z0-9_-]', '_', obj.name.strip())
                            safe_whatsapp = re.sub(r'[^0-9]', '', obj.whatsapp_number)
                            filename = f"{safe_name}_{safe_whatsapp}{ext}"
                            with open(img_path, 'rb') as img_file:
                                img_data = img_file.read()
                            zip_file.writestr(filename, img_data)
                    except Exception as e:
                        # Log error or pass silently
                        pass
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename=advance_booklet_payment_screenshots.zip'
        return response

from ckeditor.widgets import CKEditorWidget


# Removed .mpo mimetype registration to avoid integration of .mpo files
# Skipping .mpo files in export and download actions instead

@admin.action(description='Export selected biodata records to Excel with embedded photographs')
def export_to_excel(modeladmin, request, queryset):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="biodata_records_with_photos.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active

    headers = ['serial_number']
    for field in CandidateBiodata._meta.fields:
        if field.name != 'id':
            if field.name == 'education_details':
                headers.append('Education Details')  # Use friendly name
            else:
                headers.append(field.name)
    headers.append('image_found')
    ws.append(headers)

    photo_col_idx = headers.index('photograph') + 1
    ws.column_dimensions[openpyxl.utils.get_column_letter(photo_col_idx)].width = 20
    image_found_col_idx = headers.index('image_found') + 1
    ws.column_dimensions[openpyxl.utils.get_column_letter(image_found_col_idx)].width = 15

    row_num = 2
     # Get all PKs in order by submission ASC (serial 1 = oldest)
    all_ids = list(CandidateBiodata.objects.order_by('submitted_at').values_list('pk', flat=True))
    pk_to_serial = {pk: idx + 1 for idx, pk in enumerate(all_ids)}  # serial 1 = oldest

    # Filter out objects with .mpo images to avoid errors
    filtered_queryset = []
    for obj in queryset:
        photo_field = getattr(obj, 'photograph')
        if photo_field:
            try:
                img_path = photo_field.path
                ext = os.path.splitext(img_path)[1].lower()
                if ext == '.mpo':
                    logging.info(f"Excluding object with .mpo image: {img_path}")
                    continue
            except Exception as e:
                logging.error(f"Error checking image file extension: {e}")
        filtered_queryset.append(obj)
        # Order filtered_queryset by submitted_at ASC (oldest first)
    filtered_queryset = sorted(
        filtered_queryset,
        key=lambda obj: getattr(obj, 'submitted_at', None) or datetime.min,
        reverse=False
    )

    for obj in filtered_queryset:
        row = []
        # Serial number as first column
        serial_number = pk_to_serial.get(obj.pk, '-')
        row.append(serial_number)
        image_found = 'No'
        for field in CandidateBiodata._meta.fields:
            if field.name != 'id':
                if field.name == 'photograph':
                    row.append('')
                else:
                    value = getattr(obj, field.name)
                    row.append(str(value) if value is not None else '')

        photo_field = getattr(obj, 'photograph')
        img_url = None
        if photo_field:
            try:
                # Use Cloudinary URL if available
                img_url = photo_field.url
                # Fix relative URLs by prepending domain or base URL
                if img_url.startswith('/'):
                    # Assuming MEDIA_URL is /media/, prepend full domain or base URL
                    base_url = 'https://bhudevnetwork.pythonanywhere.com/'  # Replace with your actual domain or base URL
                    img_url = base_url + img_url
                ext = os.path.splitext(img_url)[1].lower()
                if ext == '.mpo':
                    logging.info(f"Skipping unsupported file format for row {row_num}: {img_url}")
                else:
                    image_found = 'Yes'
            except Exception as e:
                image_found = 'No'
                logging.error(f"Error checking image URL for row {row_num}: {e}")
        # In your export_to_excel, replace the image embedding code with:
        # if photo_field:
        #     try:
        #         img_url = photo_field.url
        #         # (your logic for constructing img_url)
        #     except Exception:
        #         img_url = ''
        #     # Instead of embedding, just put the URL in a column
        #     row.append(img_url)
        # else:
        #     row.append('')

        row.append(image_found)
        ws.append(row)

        if photo_field and image_found == 'Yes' and img_url:
            try:
                # Download image from Cloudinary URL
                logging.info(f"Downloading image for row {row_num} from URL: {img_url}")
                response_img = requests.get(img_url)
                response_img.raise_for_status()
                logging.info(f"Image download successful for row {row_num}")
                img_data = BytesIO(response_img.content)
                pil_img = PILImage.open(img_data)
                # Save to BytesIO in a format openpyxl supports (e.g., PNG)
                img_byte_arr = BytesIO()
                pil_img.save(img_byte_arr, format='PNG')
                img_byte_arr.seek(0)
                img = OpenpyxlImage(img_byte_arr)
                img.width = 80
                img.height = 80
                img.anchor = f"{openpyxl.utils.get_column_letter(photo_col_idx)}{row_num}"
                ws.add_image(img)
                ws.row_dimensions[row_num].height = 60
            except Exception as e:
                logging.error(f"Failed to embed image for row {row_num}: {e}")

        row_num += 1

    try:
        wb.save(response)
    except Exception as e:
        logging.error(f"Error saving Excel file: {e}")
        # Removed fallback save without images to ensure export always tries with images
        # Consider using Cloudinary for image hosting if local images cause issues
        # You can integrate Cloudinary URLs in export if needed
        # For now, just raise the error to notify user
        raise e

    return response

@admin.action(description='Download selected candidate images as zip')
def download_selected_images(modeladmin, request, queryset):
    zip_buffer = io.BytesIO()
    # Get all primary keys in order by submission DESC (latest first)
    all_ids = list(CandidateBiodata.objects.order_by('-submitted_at').values_list('pk', flat=True))
    total = len(all_ids)

    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        ordered_queryset = queryset.order_by('id')
        for obj in ordered_queryset:
            photo_field = getattr(obj, 'photograph')
            if photo_field and photo_field.name:
                try:
                    img_path = photo_field.path
                    ext = os.path.splitext(img_path)[1].lower()
                    if os.path.exists(img_path) and ext != '.mpo':
                        candidate_name = re.sub(r'[^a-zA-Z0-9_-]', '_', obj.candidate_name.strip())
                        mobile_number = re.sub(r'[^0-9]', '', obj.registrant_mobile or '')
                        dob_str = str(obj.dob) if obj.dob else 'unknownDOB'
                        dob_str = re.sub(r'[\\/:"*?<>|]+', '-', dob_str) 
                        try:
                            position = all_ids.index(obj.pk)
                            serial_number = total - position
                        except ValueError:
                            serial_number = "-"
                        filename = f"{serial_number}_{candidate_name}_{dob_str}_{mobile_number}{ext}"
                        with open(img_path, 'rb') as img_file:
                            img_data = img_file.read()
                        zip_file.writestr(filename, img_data)
                except Exception as e:
                    logging.error(f"Failed to add image for candidate {obj.candidate_name}: {e}")
    zip_buffer.seek(0)
    response = HttpResponse(zip_buffer, content_type='application/zip')
    zip_filename = f"selected_candidate_images_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
    response['Content-Disposition'] = f'attachment; filename={zip_filename}'
    return response
class GalleryImageAdminForm(forms.ModelForm):
    description = forms.CharField(widget=CKEditorWidget(config_name='custom'))

    class Meta:
        model = GalleryImage
        fields = '__all__'

@admin.register(CandidateBiodata)
class CandidateBiodataAdmin(admin.ModelAdmin):
    list_max_show_all = 1000 
    list_per_page = 500
    search_fields = ['candidate_name', 'registrant_mobile', 'registration_by', 'dob']
    list_filter = ['gender']

    def get_list_display(self, request):
        fields = [field.name for field in CandidateBiodata._meta.fields if field.name not in ('id', 'visa_status_details')]
        if 'visa_status' not in fields:
            visa_index = fields.index('visa_status')
            list_display = fields[:visa_index] + ['visa_status'] + fields[visa_index+1:]
        else:
            list_display = fields

        # Insert 'education_details' after 'education', only if not already there
        if 'education' in list_display and 'education_details' in list_display:
            # Remove education_details if it's before education
            list_display.remove('education_details')
            idx = list_display.index('education') + 1
            list_display = list_display[:idx] + ['education_details'] + list_display[idx:]
        elif 'education' in list_display:
            idx = list_display.index('education') + 1
            list_display = list_display[:idx] + ['education_details'] + list_display[idx:]
        elif 'education_details' not in list_display:
            list_display.append('education_details')
        new_list_display = []
        
        for field_name in list_display:
            if field_name == 'kuldevi':
                new_list_display.append('any_disability_details')
            else:
                new_list_display.append(field_name)
        return ['serial_number','duplicate_flag'] + new_list_display

    actions = [export_to_excel, download_selected_images]

    def any_disability_details(self, obj):
        return obj.kuldevi
    any_disability_details.short_description = 'Any Disability/Details'

     # --- Duplicate detection logic begins here ---
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        # Count occurrences of (candidate_name, dob)
        from collections import defaultdict
        seen = defaultdict(list)
        for obj in qs:
            key = (obj.candidate_name, obj.dob)
            seen[key].append(obj.pk)
        # Store duplicate indices for quick access
        self.duplicate_indices = set()
        for key, ids in seen.items():
            if len(ids) > 1:
                # Mark all except the first as duplicate
                self.duplicate_indices.update(ids[1:])
        return qs

    def duplicate_flag(self, obj):
        if hasattr(self, 'duplicate_indices') and obj.pk in self.duplicate_indices:
            # You can return HTML with a style or icon, or just a flag for CSS
            return format_html('<span style="background-color: #ffcccc; color: red; font-weight: bold;">DUPLICATE</span>')
        return ""
    duplicate_flag.short_description = "Duplicate?"

    def serial_number(self, obj):
        # Get all primary keys in order by submission DESC (latest first)
        qs = CandidateBiodata.objects.order_by('-submitted_at').values_list('pk', flat=True)
        total = qs.count()
        try:
            position = list(qs).index(obj.pk)
            return total - position
        except ValueError:
            return '-'
    serial_number.short_description = 'Serial Number'
    def education_details(self, obj):
        # Put any logic you want here, or simply display a placeholder
        return obj.education_details  # Or: return obj.education, or any logic you want
    education_details.short_description = 'Education Details'

@admin.register(GalleryImage)
class GalleryImageAdmin(admin.ModelAdmin):
    form = GalleryImageAdminForm
    list_display = ['title', 'uploaded_at', 'description_display']

    def description_display(self, obj):
        return format_html('<div style="font-size:14px;">{}</div>', obj.description)
    description_display.short_description = 'Description'
from django.utils.safestring import mark_safe

@admin.register(AdvancePassBooking)
class AdvancePassBookingAdmin(admin.ModelAdmin):
    list_display = ['name', 'city', 'whatsapp_number', 'email', 'entry_token_quantity', 'tea_coffee_quantity', 'unlimited_buffet_quantity', 'total_amount', 'payment_screenshot_preview', 'created_at']
    list_filter = ['created_at']
    search_fields = ['name', 'city', 'whatsapp_number', 'email']

    readonly_fields = ['payment_screenshot_preview']

    actions = ['export_selected_to_excel', 'export_payment_screenshots_zip']

    def payment_screenshot_preview(self, obj):
        if obj.payment_screenshot:
            return mark_safe(f'<img src="{obj.payment_screenshot.url}" style="max-height: 100px; max-width: 100px;" />')
        return "-"
    payment_screenshot_preview.short_description = 'Payment Screenshot'

    @admin.action(description='Export selected advance pass bookings to Excel')
    def export_selected_to_excel(self, request, queryset):
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as OpenpyxlImage
        from io import BytesIO
        from django.http import HttpResponse
        import requests
        from PIL import Image as PILImage

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Advance Pass Bookings"

        # Define headers
        headers = ['Name', 'City', 'WhatsApp Number', 'Email', 'Pass Type', 'Quantity', 'Total Amount', 'Payment Screenshot', 'Created At']
        ws.append(headers)

        # Set column widths
        column_widths = [20, 20, 20, 30, 15, 10, 15, 20, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        row_num = 2
        for obj in queryset:
            row = [
                obj.name,
                obj.city,
                obj.whatsapp_number,
                obj.email,
                obj.pass_type,
                obj.quantity,
                obj.total_amount,
                '',  # Placeholder for image
                obj.created_at.strftime('%Y-%m-%d %H:%M:%S') if obj.created_at else '',
            ]
            ws.append(row)

            # Embed payment screenshot image if available
            if obj.payment_screenshot:
                try:
                    img_url = obj.payment_screenshot.url
                    if img_url.startswith('/'):
                        base_url = 'https://bhudevnetwork.pythonanywhere.com/'  # Replace with your actual domain or base URL
                        img_url = base_url + img_url
                    response_img = requests.get(img_url)
                    response_img.raise_for_status()
                    img_data = BytesIO(response_img.content)
                    pil_img = PILImage.open(img_data)
                    img_byte_arr = BytesIO()
                    pil_img.save(img_byte_arr, format='PNG')
                    img_byte_arr.seek(0)
                    img = OpenpyxlImage(img_byte_arr)
                    img.width = 80
                    img.height = 80
                    img.anchor = f"{get_column_letter(8)}{row_num}"
                    ws.add_image(img)
                    ws.row_dimensions[row_num].height = 60
                except Exception as e:
                    # Log error or pass silently
                    pass
            row_num += 1

        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=advance_pass_bookings.xlsx'

        # Save workbook to response
        wb.save(response)
        return response

    @admin.action(description='Export payment screenshots of selected advance pass bookings as ZIP')
    def export_payment_screenshots_zip(self, request, queryset):
        import zipfile
        import os
        from django.http import HttpResponse
        from io import BytesIO
        import re

        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for obj in queryset:
                if obj.payment_screenshot:
                    try:
                        img_path = obj.payment_screenshot.path
                        ext = os.path.splitext(img_path)[1].lower()
                        if os.path.exists(img_path):
                            # Sanitize name and whatsapp number for filename
                            safe_name = re.sub(r'[^a-zA-Z0-9_-]', '_', obj.name.strip())
                            safe_whatsapp = re.sub(r'[^0-9]', '', obj.whatsapp_number)
                            filename = f"{safe_name}_{safe_whatsapp}{ext}"
                            with open(img_path, 'rb') as img_file:
                                img_data = img_file.read()
                            zip_file.writestr(filename, img_data)
                    except Exception as e:
                        # Log error or pass silently
                        pass
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename=payment_screenshots.zip'
        return response
